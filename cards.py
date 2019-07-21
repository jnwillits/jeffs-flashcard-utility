#!/usr/bin/python

'''
Jeff's Flash Card System  - Jeffrey Neil Willits  @jnwillits
'''

import copy
import json
import os
from random import shuffle
import sys
from math import ceil

import openpyxl
import PySimpleGUI as sg


sg.ChangeLookAndFeel('Dark')
sg.SetOptions(icon='cards.ico', element_padding=(5, 0), font=('verdana', 10), text_color='#32CD32',
              background_color='#1E1E1E', text_element_background_color='#1E1E1E')


menu_def = [['Setup', ['Reset', 'Quit']],
            ['Help', 'About...'],
            ]

l_pad = 65
frame_color = '#2a2a2a'

f_frame_layout = [
    [sg.T('', background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L0_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L1_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L2_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L3_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L4_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L5_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L6_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L7_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L8_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L9_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L10_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L11_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L12_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L13_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L14_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L15_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L16_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L17_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L18_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L19_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_F_L20_',background_color=frame_color)],
]

b_frame_layout = [
    [sg.T('', background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L0_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L1_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L2_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L3_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L4_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L5_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L6_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L7_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L8_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L9_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L10_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L11_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L12_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L13_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L14_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L15_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L16_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L17_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L18_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L19_',background_color=frame_color)],
    [sg.T('', size=(l_pad, 0), key='_B_L20_',background_color=frame_color)],
]

layout = [
    [sg.Menu(menu_def, tearoff=False, pad=(20, 1))],
    [sg.T('')],
    [sg.T('', text_color='#FFFAFA', size=(20, 0), key='_status1_'),
     sg.T('', text_color='#FFFAFA', size=(20, 0), key='_status2_')],
    # [sg.Text('_' * 127, text_color='#565656', size=(0, 1))],
    [sg.T('')],
    [sg.Frame('Front', f_frame_layout, border_width=0, background_color=frame_color),sg.T('', size=(1, 0)),
     sg.Frame('Back', b_frame_layout, border_width=0, background_color=frame_color),],
    [sg.T('')],
    [sg.Button('', visible=False, size=(60, 30), ), ],
    [sg.T('', size=(95, 0)), sg.Button('Archive', button_color=('#FFFFFF', '#565656'), visible=True, pad=(0, 0),
        size=(12, 1)), sg.Button('Next', button_color=('#FFFFFF', '#565656'), visible=True, pad=(40, 0), size=(12, 1))], ]


def save_data_and_exit(active_cards_pass):
    with open('cards.json', 'w') as f_obj:
        json.dump(active_cards_pass, f_obj)
    window.Close()
    sys.exit()


def populate_data(ws_pass, col_pass, rows_per_card_pass, cards_pass, card_side):
    card_num = 1
    card_line = 1
    for row in ws_pass[col_pass]:
        if row.value == None:
            cell_val = ''
        else:
            cell_val = row.value
        cards_pass[card_num - 1][card_side][card_line - 1] = cell_val
        if card_line < rows_per_card_pass:
            card_line += 1
        else:
            if card_num < num_cards:
                card_line = 1
                card_num += 1
    return cards_pass


if __name__ == '__main__':
    rows_per_card = 20
    wb = openpyxl.load_workbook('cards.xlsx', data_only=True)
    ws = wb['Sheet1']

    # The last card may have less than rows_per_card. This assures num_cards accounts for the last card.
    max_rows = max(len(ws['B']), len(ws['D']))
    if max_rows % rows_per_card == 0:
        num_cards=max_rows // rows_per_card
    else:
        num_cards=ceil(max_rows / rows_per_card)

    card=[]
    card.append(['' for element in range(rows_per_card)])
    card.append(['' for element in range(rows_per_card)])
    cards=[copy.deepcopy(card) for element in range(num_cards)]

    cards=populate_data(ws, 'B', rows_per_card, cards, card_side=0)
    cards=populate_data(ws, 'D', rows_per_card, cards, card_side=1)

    window=sg.Window(" Jeff's Flash Card System", size=(1200, 600), default_element_size=(30, 1), grab_anywhere=False,
                       background_color='#1E1E1E', auto_size_text=False, auto_size_buttons=False).Layout(layout).Finalize()

    if os.path.isfile('cards.json'):
        with open('cards.json') as f_obj:
            active_cards=json.load(f_obj)
    else:
        active_cards=list(i for i in range(num_cards))

    same_card_num=0
    show_front=True
    while True:
        event, values=window.Read(timeout=10)
        window.Element('_status1_').Update(f'Cards in deck: {num_cards}')
        window.Element('_status2_').Update(
            f'      Active cards: {len(active_cards)}')
        if event is None or event == 'Exit':
            break
        else:
            if event == 'About...':
                sg.PopupNoButtons('', 'This is a flash card study utility that',
                                  'uses an Excel spreadsheet as the database.',
                                  'Set up your own deck by editing the spreadsheet.\n',
                                  'The Python source code is available from my',
                                  'GitHub repository.\n',
                                  'Version 1.1 released July 19, 2019.\n\n'
                                  'Jeffrey Neil Willits', '@jnwillits\n', no_titlebar=False, keep_on_top=True,
                                  grab_anywhere=True, background_color='#000000')
            elif event == 'Reset':
                active_cards=list(i for i in range(num_cards))
            elif event == 'Archive':
                current_card=active_cards[0]
                if len(active_cards) >= 1:
                    active_cards.remove(current_card)
                else:
                    sg.Popup('', 'All cards are archived. Use the menu to reset the deck',
                                 'to see them again.\n', no_titlebar=False, keep_on_top=True, grab_anywhere=True)
            elif event == 'Next':
                if len(active_cards) >= 1:
                    if show_front:
                        while True:
                            shuffle(active_cards)
                            current_card=active_cards[0]
                            if same_card_num == current_card:
                                continue
                            else:
                                same_card_num=current_card
                                break
                        for row in range(rows_per_card):
                            window.Element(
                                '_F_L' + str(row) + '_').Update(f'{cards[current_card][0][row]}')
                            window.Element('_B_L' + str(row) + '_').Update('')
                            show_front=False
                    else:
                        for row in range(rows_per_card):
                            window.Element(
                                '_F_L' + str(row) + '_').Update(f'{cards[current_card][0][row]}')
                            window.Element(
                                '_B_L' + str(row) + '_').Update(f'{cards[current_card][1][row]}')
                            show_front=True
                else:
                    sg.Popup('', 'All cards are archived. Use the menu to reset the deck',
                                 'to see them again.\n', no_titlebar=False, keep_on_top=True, grab_anywhere=True)
            elif event == 'Quit':
                save_data_and_exit(active_cards)

save_data_and_exit(active_cards)
